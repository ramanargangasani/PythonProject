import openpyxl

book = openpyxl.load_workbook("C:\\Users\\HarshaLalith\\Desktop\\Python Info\\exceltestdata.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
Dict = {}
'''
# set the values into the excel sheet
sheet.cell(row=4, column=2).value = "Aditya"
sheet.cell(row=5, column=3).value = "Gangasani"
sheet.cell(row=6, column=4).value = "Aditya.g@gmail.com"
'''

# max column count and row count
print(sheet.max_row)
print(sheet.max_column)
# perticular cell value
print(sheet['A3'].value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column + 1):
            #
            Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)


