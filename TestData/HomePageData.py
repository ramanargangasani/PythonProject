import openpyxl

class HomePageData:
    test_HomePage_Data = [{"firstname":"Ramana", "lastname":"Gangasani", "gender" : "Male"}, {"firstname":"RamaDevi", "lastname":"Gangasani", "gender" : "Female"}]

    @staticmethod                   # if decalare staticmethod there is no self. if self required you need to declare a object
    def getExcelTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\HarshaLalith\\Desktop\\Python Info\\exceltestdata.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:    # test_case_name = Testcase2
                for j in range(2, sheet.max_column + 1):
                    # Dict['lastname'] ="Gangasani"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]

    @staticmethod  # if decalare staticmethod there is no self. if self required you need to declare a object
    def getExcelTestData2(test_case_name):  # this method for testing git hub should be removed after that
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\HarshaLalith\\Desktop\\Python Info\\exceltestdata.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:  # test_case_name = Testcase2
                for j in range(2, sheet.max_column + 1):
                    # Dict['lastname'] ="Ramana"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]




